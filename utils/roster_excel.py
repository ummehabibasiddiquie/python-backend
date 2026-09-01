"""
Weekly roster Excel template + parse helpers.

Managers download a week grid (Team Member × Mon–Sun), edit via dropdowns,
then upload. Parsed cells become DAY_UPDATE / LEAVE_ADD change requests.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.roster_helpers import (
    AGENT_DAY_SHIFT_END,
    AGENT_DAY_SHIFT_START,
    QA_DAY_SHIFT_END,
    QA_DAY_SHIFT_START,
    month_year_label,
    parse_date,
)

# Display labels — keep aligned with managers' existing Excel format
LABEL_AGENT_DAY = "9:00AM to 6:30PM"
LABEL_QA_DAY = "10:00AM to 7:30PM"
LABEL_NIGHT = "7:30 PM to 8:30 AM"
LABEL_WEEK_OFF = "Week Off"
LABEL_LEAVE = "Leave"
LABEL_LEAVE_AFFECT_TARGET = "Leave (Affect Target)"
LABEL_HALF_DAY = "Half day"
LABEL_HALF_DAY_AFFECT_TARGET = "Half day (Affect Target)"

DROPDOWN_VALUES = [
    LABEL_AGENT_DAY,
    LABEL_QA_DAY,
    LABEL_NIGHT,
    LABEL_WEEK_OFF,
    LABEL_LEAVE,
    LABEL_LEAVE_AFFECT_TARGET,
    LABEL_HALF_DAY,
    LABEL_HALF_DAY_AFFECT_TARGET,
]

NIGHT_SHIFT_START = time(19, 30)
NIGHT_SHIFT_END = time(8, 30)

DAY_NAME_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

_HEADER_RE = re.compile(
    r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s*\((\d{1,2})[-/](\d{1,2})[-/](\d{4})\)$",
    re.IGNORECASE,
)


def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def week_dates(week_start: date) -> list[date]:
    start = monday_of_week(week_start)
    return [start + timedelta(days=i) for i in range(7)]


def weeks_in_month(year: int, month: int) -> list[dict]:
    """
    Calendar weeks (Mon–Sun) that touch this month, numbered Week 1..N
    for manager-friendly picker / full-month week-by-week upload.
    """
    first = date(year, month, 1)
    if month == 12:
        last = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(year, month + 1, 1) - timedelta(days=1)

    cursor = monday_of_week(first)
    weeks: list[dict] = []
    seen: set[date] = set()

    while cursor <= last:
        if cursor not in seen:
            days = week_dates(cursor)
            if any(d.month == month and d.year == year for d in days):
                weeks.append(
                    {
                        "week_start": cursor.isoformat(),
                        "week_end": days[-1].isoformat(),
                        "days_in_month": sum(
                            1 for d in days if d.month == month and d.year == year
                        ),
                        "date_range": f"{days[0].strftime('%d %b')} - {days[-1].strftime('%d %b %Y')}",
                    }
                )
                seen.add(cursor)
        cursor += timedelta(days=7)

    for i, w in enumerate(weeks, start=1):
        w["week_number"] = i
        w["week_key"] = f"week_{i}"
        w["label"] = f"Week {i} ({w['date_range']})"
        w["short_label"] = f"Week {i}"
    return weeks


def sheet_title_for_week(week_number: int) -> str:
    return f"Week {int(week_number)}"


def is_roster_data_sheet(name: str) -> bool:
    n = (name or "").strip().lower()
    if n in ("instructions", "_lists"):
        return False
    if n == "roster":
        return True
    return bool(re.match(r"^week\s*\d+$", n))


def _normalize_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and str(value) == "nan":
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    # Collapse whitespace / common variants
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").strip().lower())


def parse_header_date(header: Any) -> date | None:
    """Parse 'Mon (27-07-2026)' or similar header cells."""
    if isinstance(header, datetime):
        return header.date()
    if isinstance(header, date):
        return header
    s = _normalize_label(header)
    if not s:
        return None
    m = _HEADER_RE.match(s)
    if m:
        day, month, year = int(m.group(2)), int(m.group(3)), int(m.group(4))
        return date(year, month, day)
    # Fallback: try common date formats
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def format_day_header(d: date) -> str:
    return f"{DAY_NAME_ABBR[d.weekday()]} ({d.strftime('%d-%m-%Y')})"


def _coerce_time(value: Any) -> time | None:
    """Normalize DB/API time values (time, datetime, timedelta, str) to datetime.time."""
    if value is None:
        return None
    if isinstance(value, time) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.time()
    # mysql-connector returns TIME columns as timedelta
    if isinstance(value, timedelta):
        total = int(value.total_seconds()) % (24 * 3600)
        if total < 0:
            total += 24 * 3600
        hours, rem = divmod(total, 3600)
        minutes, seconds = divmod(rem, 60)
        return time(hours, minutes, seconds)
    s = str(value).strip()
    if not s or s.lower() in ("none", "null", "nat"):
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s[: len(fmt) + 2], fmt).time()
        except ValueError:
            continue
    return None


def _time_close(a: time | None, b: time, tol_minutes: int = 15) -> bool:
    if a is None:
        return False
    mins_a = a.hour * 60 + a.minute
    mins_b = b.hour * 60 + b.minute
    return abs(mins_a - mins_b) <= tol_minutes


def day_to_excel_label(day: dict | None, role_name: str | None = None) -> str:
    """Map a roster_day row to a dropdown label for template prefill."""
    if not day:
        return ""
    day_type = (day.get("day_type") or "").strip()
    if day_type == "WeekOff":
        return LABEL_WEEK_OFF
    if day_type == "Leave":
        working_type = (day.get("working_type") or "Full").strip()
        is_half = working_type == "Half" or bool(int(day.get("leave_is_half_day") or 0))
        affect = day.get("leave_affect_target")
        if affect is None:
            affect = day.get("affect_target")
        affect = bool(int(affect or 0))
        if is_half:
            return LABEL_HALF_DAY_AFFECT_TARGET if affect else LABEL_HALF_DAY
        if affect:
            return LABEL_LEAVE_AFFECT_TARGET
        return LABEL_LEAVE
    if day_type == "Holiday":
        return LABEL_WEEK_OFF  # managers usually don't set Holiday in weekly Excel
    if day_type == "PreJoin":
        return ""
    if day_type != "Working":
        return ""

    working_type = (day.get("working_type") or "Full").strip()
    if working_type == "Half":
        # Working half-day always reduces hours/days (same as half leave that affects target).
        return LABEL_HALF_DAY_AFFECT_TARGET

    shift = (day.get("shift") or "DAY").strip().upper()
    if shift == "NIGHT":
        return LABEL_NIGHT

    start = _coerce_time(day.get("shift_start"))

    if _time_close(start, AGENT_DAY_SHIFT_START):
        return LABEL_AGENT_DAY
    if _time_close(start, QA_DAY_SHIFT_START):
        return LABEL_QA_DAY
    if (role_name or "").strip().lower() == "qa":
        return LABEL_QA_DAY
    return LABEL_AGENT_DAY


def excel_label_to_change(
    label: str,
    roster_date: date,
    *,
    role_name: str | None = None,
) -> dict | None:
    """
    Convert a cell label into a change request descriptor.
    Returns None for empty cells.
    Raises ValueError for unrecognized values.
    """
    raw = _normalize_label(label)
    if not raw:
        return None

    key = _normalize_key(raw)
    date_str = roster_date.isoformat()

    # Aliases
    alias_map = {
        _normalize_key(LABEL_AGENT_DAY): "agent_day",
        _normalize_key(LABEL_QA_DAY): "qa_day",
        _normalize_key(LABEL_NIGHT): "night",
        _normalize_key(LABEL_WEEK_OFF): "week_off",
        _normalize_key(LABEL_LEAVE_AFFECT_TARGET): "leave_affect_target",
        _normalize_key(LABEL_LEAVE): "leave",
        _normalize_key(LABEL_HALF_DAY_AFFECT_TARGET): "half_day_affect_target",
        _normalize_key(LABEL_HALF_DAY): "half_day",
        "weekoff": "week_off",
        "wo": "week_off",
        "off": "week_off",
        "halfdayaffecttarget": "half_day_affect_target",
        "halfday": "half_day",
        "half": "half_day",
        "nighthift": "night",
        "night": "night",
        "dayshift": "agent_day",
        "day": "agent_day",
        "leaveaffecttarget": "leave_affect_target",
        "leaveaffectsarget": "leave_affect_target",
        "leavewithtarget": "leave_affect_target",
    }
    kind = alias_map.get(key)
    if not kind:
        # Fuzzy: contains
        lower = raw.lower()
        if "week" in lower and "off" in lower:
            kind = "week_off"
        elif "half" in lower and ("affect" in lower or "target" in lower):
            kind = "half_day_affect_target"
        elif "half" in lower:
            kind = "half_day"
        elif "leave" in lower and ("affect" in lower or "target" in lower):
            kind = "leave_affect_target"
        elif "leave" in lower:
            kind = "leave"
        elif "7:30" in lower.replace(" ", "") and ("pm" in lower or "p.m" in lower):
            kind = "night"
        elif "10:00" in lower.replace(" ", "") or "10am" in lower.replace(" ", ""):
            kind = "qa_day"
        elif "9:00" in lower.replace(" ", "") or "9am" in lower.replace(" ", ""):
            kind = "agent_day"
        else:
            raise ValueError(f"Unrecognized value: {raw}")

    if kind == "week_off":
        return {
            "change_type": "DAY_UPDATE",
            "label": LABEL_WEEK_OFF,
            "change_payload": {
                "roster_date": date_str,
                "day_type": "WeekOff",
                "shift": "DAY",
                "working_type": "Full",
                "working_hours": 0,
            },
        }

    if kind in ("leave", "leave_affect_target", "half_day", "half_day_affect_target"):
        is_half = 1 if kind in ("half_day", "half_day_affect_target") else 0
        affect_target = 1 if kind in ("leave_affect_target", "half_day_affect_target") else 0
        if is_half:
            label = LABEL_HALF_DAY_AFFECT_TARGET if affect_target else LABEL_HALF_DAY
        else:
            label = LABEL_LEAVE_AFFECT_TARGET if affect_target else LABEL_LEAVE
        return {
            "change_type": "LEAVE_ADD",
            "label": label,
            "change_payload": {
                "leave_type": "Leave",
                "start_date": date_str,
                "end_date": date_str,
                "reason": "Roster Excel upload",
                "affect_target": affect_target,
                "is_half_day": is_half,
                "is_rostered": 1,
            },
        }

    if kind == "night":
        return {
            "change_type": "DAY_UPDATE",
            "label": LABEL_NIGHT,
            "change_payload": {
                "roster_date": date_str,
                "day_type": "Working",
                "shift": "NIGHT",
                "working_type": "Full",
                "shift_start": NIGHT_SHIFT_START.strftime("%H:%M:%S"),
                "shift_end": NIGHT_SHIFT_END.strftime("%H:%M:%S"),
            },
        }

    # Day shifts — cell value wins (explicit 9am vs 10am)
    if kind == "qa_day":
        start, end = QA_DAY_SHIFT_START, QA_DAY_SHIFT_END
        label_out = LABEL_QA_DAY
    else:
        start, end = AGENT_DAY_SHIFT_START, AGENT_DAY_SHIFT_END
        label_out = LABEL_AGENT_DAY

    return {
        "change_type": "DAY_UPDATE",
        "label": label_out,
        "change_payload": {
            "roster_date": date_str,
            "day_type": "Working",
            "shift": "DAY",
            "working_type": "Full",
            "shift_start": start.strftime("%H:%M:%S"),
            "shift_end": end.strftime("%H:%M:%S"),
        },
    }


def _current_day_signature(day: dict | None) -> tuple:
    if not day:
        return ("", "", "", "")
    return (
        (day.get("day_type") or "").strip(),
        (day.get("shift") or "").strip().upper(),
        (day.get("working_type") or "Full").strip(),
        day_to_excel_label(day),
    )


def proposed_signature(change: dict) -> tuple:
    ct = change.get("change_type")
    p = change.get("change_payload") or {}
    if ct == "LEAVE_ADD":
        half = int(p.get("is_half_day") or 0)
        affect = int(p.get("affect_target") or 0)
        if half:
            leave_label = LABEL_HALF_DAY_AFFECT_TARGET if affect else LABEL_HALF_DAY
        else:
            leave_label = LABEL_LEAVE_AFFECT_TARGET if affect else LABEL_LEAVE
        return ("Leave", "DAY", "Half" if half else "Full", leave_label, affect)
    return (
        (p.get("day_type") or "").strip(),
        (p.get("shift") or "DAY").strip().upper(),
        (p.get("working_type") or "Full").strip(),
        change.get("label") or "",
        0,
    )


def is_noop_change(day: dict | None, change: dict) -> bool:
    """Skip creating a request when Excel matches current roster state."""
    if not day:
        return False
    cur_type, cur_shift, cur_wt, cur_label = _current_day_signature(day)
    prop = proposed_signature(change)
    prop_type, prop_shift, prop_wt, prop_label = prop[0], prop[1], prop[2], prop[3]

    if change.get("change_type") == "LEAVE_ADD":
        if cur_type != "Leave":
            return False
        payload = change.get("change_payload") or {}
        prop_half = int(payload.get("is_half_day") or 0)
        prop_affect = int(payload.get("affect_target") or 0)
        cur_half = 1 if (
            cur_wt == "Half"
            or bool(int(day.get("leave_is_half_day") or 0))
        ) else 0
        if cur_half != prop_half:
            return False
        cur_affect = int(
            day.get("leave_affect_target")
            if day.get("leave_affect_target") is not None
            else day.get("affect_target") or 0
        )
        return cur_affect == prop_affect

    if prop_type == "WeekOff":
        return cur_type == "WeekOff"
    if prop_type == "Working" and prop_wt == "Half":
        return cur_type == "Working" and cur_wt == "Half"
    if prop_type == "Working" and prop_shift == "NIGHT":
        return cur_type == "Working" and cur_shift == "NIGHT" and cur_wt != "Half"
    if prop_type == "Working":
        # Compare by label so agent/QA day times are respected
        return (
            cur_type == "Working"
            and cur_shift == "DAY"
            and cur_wt != "Half"
            and cur_label == prop_label
        )
    return False


def _add_week_sheet(
    wb: Workbook,
    *,
    title: str,
    week_start: date,
    employees: list[dict],
    day_lookup: dict[tuple[int, str], dict],
) -> None:
    days = week_dates(week_start)
    ws = wb.create_sheet(title)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    headers = ["Team Member"] + [format_day_header(d) for d in days]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    ws.column_dimensions["A"].width = 28
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22

    for row_idx, emp in enumerate(employees, start=2):
        name = emp.get("user_name") or f"User {emp.get('user_id')}"
        role = emp.get("role_name")
        ws.cell(row_idx, 1, name).border = thin
        uid = int(emp["user_id"])
        for col_idx, d in enumerate(days, start=2):
            day = day_lookup.get((uid, d.isoformat()))
            label = day_to_excel_label(day, role)
            # Defaults when roster has no value yet:
            # Sat/Sun → Week Off; Mon–Fri → role day shift
            if not label:
                if d.weekday() >= 5:
                    label = LABEL_WEEK_OFF
                elif (role or "").strip().lower() == "qa":
                    label = LABEL_QA_DAY
                else:
                    label = LABEL_AGENT_DAY
            cell = ws.cell(row_idx, col_idx, label or "")
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    last_row = max(len(employees) + 1, 2)
    extra_end = last_row + 15
    for r in range(last_row + 1, extra_end + 1):
        for c in range(1, 9):
            ws.cell(r, c).border = thin

    n = len(DROPDOWN_VALUES)
    dv = DataValidation(
        type="list",
        formula1=f"=_Lists!$A$1:$A${n}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Please select a value from the dropdown list.",
    )
    ws.add_data_validation(dv)
    dv.add(f"B2:H{extra_end}")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:H{last_row}"


def build_month_workbook(
    *,
    weeks: list[dict],
    employees: list[dict],
    day_lookup: dict[tuple[int, str], dict] | None = None,
    month_label: str = "",
) -> bytes:
    """One workbook with a sheet per week (Week 1, Week 2, …)."""
    if not weeks:
        raise ValueError("No weeks to include in month workbook")

    day_lookup = day_lookup or {}
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    info = wb.create_sheet("Instructions", 0)
    info["A1"] = "Monthly Roster Upload — Instructions"
    info["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        f"Month: {month_label}" if month_label else "",
        "This file has one sheet per week (Week 1, Week 2, Week 3, …).",
        "1. Do not rename week sheets or change header row formats.",
        "2. Column A (Team Member) is prefilled — names must match HRMS.",
        "3. Use the dropdown in each day cell.",
        "4. Allowed values:",
        f"   - {LABEL_AGENT_DAY}",
        f"   - {LABEL_QA_DAY}",
        f"   - {LABEL_NIGHT}",
        f"   - {LABEL_WEEK_OFF}",
        f"   - {LABEL_LEAVE}  (does NOT affect monthly target)",
        f"   - {LABEL_LEAVE_AFFECT_TARGET}  (DOES reduce monthly target)",
        f"   - {LABEL_HALF_DAY}  (works half day; does NOT affect monthly target)",
        f"   - {LABEL_HALF_DAY_AFFECT_TARGET}  (works half day; DOES reduce monthly target by 0.5)",
        "5. Fill Week 1, Week 2, … as needed (later weeks can stay blank for a later upload).",
        "6. Upload from Roster Management → Excel Upload. All week sheets are read together.",
        "7. You can also download / upload a single week anytime.",
        "",
        "Weeks in this file:",
    ]
    for w in weeks:
        lines.append(f"   - {w.get('label') or sheet_title_for_week(w.get('week_number') or 0)}")
    for i, line in enumerate(lines, start=2):
        info[f"A{i}"] = line
    info.column_dimensions["A"].width = 100

    lists = wb.create_sheet("_Lists")
    for i, val in enumerate(DROPDOWN_VALUES, start=1):
        lists[f"A{i}"] = val
    lists.sheet_state = "hidden"

    for w in weeks:
        week_start = parse_date(w["week_start"])
        if not week_start:
            continue
        title = sheet_title_for_week(int(w["week_number"]))
        _add_week_sheet(
            wb,
            title=title,
            week_start=week_start,
            employees=employees,
            day_lookup=day_lookup,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_template_workbook(
    *,
    week_start: date,
    employees: list[dict],
    day_lookup: dict[tuple[int, str], dict] | None = None,
    week_number: int | None = None,
) -> bytes:
    """
    Build weekly roster xlsx with dropdowns.
    employees: [{user_id, user_name, role_name}, ...]
    day_lookup: (user_id, YYYY-MM-DD) -> roster_day dict
    """
    days = week_dates(week_start)
    day_lookup = day_lookup or {}

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    week_title = f"Week {week_number}" if week_number else "Weekly"
    info = wb.create_sheet("Instructions", 0)
    info["A1"] = f"{week_title} Roster Upload — Instructions"
    info["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "1. Do not rename the roster sheet or change the header row format.",
        "2. Column A (Team Member) is prefilled — do not rename agents (names must match HRMS).",
        "3. Use the dropdown in each day cell to pick a value.",
        "4. Allowed values:",
        f"   - {LABEL_AGENT_DAY}",
        f"   - {LABEL_QA_DAY}",
        f"   - {LABEL_NIGHT}",
        f"   - {LABEL_WEEK_OFF}",
        f"   - {LABEL_LEAVE}  (does NOT affect monthly target)",
        f"   - {LABEL_LEAVE_AFFECT_TARGET}  (DOES reduce monthly target)",
        f"   - {LABEL_HALF_DAY}  (works half day; does NOT affect monthly target)",
        f"   - {LABEL_HALF_DAY_AFFECT_TARGET}  (works half day; DOES reduce monthly target by 0.5)",
        "5. Save the file and upload it from Roster Management → Excel Upload.",
        "6. Review the preview, then confirm. Changes become pending requests (submit for approval as usual).",
        "7. You can still add / edit / delete days from the roster calendar after upload.",
        "8. For a full month, pick Week 1 / Week 2 / Week 3… or use Download all weeks.",
        "",
        f"Week: {format_day_header(days[0])} to {format_day_header(days[-1])}",
    ]
    for i, line in enumerate(lines, start=2):
        info[f"A{i}"] = line
    info.column_dimensions["A"].width = 100

    lists = wb.create_sheet("_Lists")
    for i, val in enumerate(DROPDOWN_VALUES, start=1):
        lists[f"A{i}"] = val
    lists.sheet_state = "hidden"

    sheet_name = sheet_title_for_week(week_number) if week_number else "Roster"
    _add_week_sheet(
        wb,
        title=sheet_name,
        week_start=week_start,
        employees=employees,
        day_lookup=day_lookup,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_week_sheet(ws) -> dict:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    date_cols: list[tuple[int, date]] = []
    for c in range(2, len(headers) + 1):
        d = parse_header_date(headers[c - 1] if c - 1 < len(headers) else None)
        if d:
            date_cols.append((c, d))

    if not date_cols:
        raise ValueError(
            f"Sheet '{ws.title}': could not find date columns. Headers must look like: Mon (27-07-2026)"
        )

    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = _normalize_label(ws.cell(r, 1).value)
        if not name:
            if not any(_normalize_label(ws.cell(r, c).value) for c, _ in date_cols):
                continue
            raise ValueError(f"Sheet '{ws.title}' row {r}: Team Member name is missing")

        cells: dict[str, str] = {}
        for c, d in date_cols:
            val = _normalize_label(ws.cell(r, c).value)
            if val:
                cells[d.isoformat()] = val
        if cells:
            rows.append({"row": r, "name": name, "cells": cells, "sheet": ws.title})

    return {
        "sheet": ws.title,
        "dates": [d for _, d in date_cols],
        "rows": rows,
    }


def parse_roster_excel(file_bytes: bytes) -> dict:
    """
    Parse uploaded weekly or multi-week roster workbook.
    Supports single 'Roster' / 'Week N' sheet, or full-month file with Week 1..N sheets.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    data_sheets = [wb[name] for name in wb.sheetnames if is_roster_data_sheet(name)]
    if not data_sheets:
        for name in wb.sheetnames:
            if name.strip().lower() not in ("instructions", "_lists"):
                data_sheets = [wb[name]]
                break
    if not data_sheets:
        raise ValueError("No roster sheet found in Excel file")

    all_dates: list[date] = []
    all_rows: list[dict] = []
    sheets_meta: list[dict] = []
    seen_dates: set[date] = set()

    for ws in data_sheets:
        parsed = _parse_week_sheet(ws)
        sheets_meta.append(
            {
                "sheet": parsed["sheet"],
                "dates": [d.isoformat() for d in parsed["dates"]],
                "row_count": len(parsed["rows"]),
            }
        )
        for d in parsed["dates"]:
            if d not in seen_dates:
                seen_dates.add(d)
                all_dates.append(d)
        all_rows.extend(parsed["rows"])

    all_dates.sort()
    return {
        "dates": all_dates,
        "rows": all_rows,
        "sheets": sheets_meta,
    }


def month_years_for_dates(dates: list[date]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for d in dates:
        label = month_year_label(d.year, d.month)
        if label not in seen:
            seen.add(label)
            out.append(label)
    return out


def match_employee_name(
    name: str,
    employees_by_norm: dict[str, list[dict]],
) -> tuple[dict | None, str | None]:
    """Return (employee, error)."""
    key = _normalize_key(name)
    if not key:
        return None, f"No matching employee for '{name}'"

    matches = employees_by_norm.get(key) or []
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, f"Ambiguous name '{name}' matches multiple employees"

    # Short Excel names (e.g. team codes "A" / "B") — exact match only.
    # Never use them in substring matching (would hit every name containing that letter).
    MIN_PARTIAL_LEN = 3
    if len(key) < MIN_PARTIAL_LEN:
        return None, f"No matching employee for '{name}'"

    # Partial contains match as fallback (longer names only).
    # Also skip short *employee* keys so a user named "A" cannot steal
    # another row via `"a" in "priyasharma"`.
    partial = []
    for k, emps in employees_by_norm.items():
        if not k or len(k) < MIN_PARTIAL_LEN:
            continue
        if key in k or k in key:
            partial.extend(emps)
    # unique by user_id
    uniq: dict[int, dict] = {int(e["user_id"]): e for e in partial}
    if len(uniq) == 1:
        return next(iter(uniq.values())), None
    if len(uniq) > 1:
        return None, f"Ambiguous name '{name}'"
    return None, f"No matching employee for '{name}'"


def resolve_roster_month_for_date(
    roster_by_user_month: dict[tuple[int, str], dict],
    user_id: int,
    d: date,
) -> dict | None:
    my = month_year_label(d.year, d.month)
    return roster_by_user_month.get((int(user_id), my))
